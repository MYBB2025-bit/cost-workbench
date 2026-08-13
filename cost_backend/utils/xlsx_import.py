"""xlsx 导入工具：把 Excel 字节解析成统一字段字典列表。

支持中文/英文表头，列映射：
- 编号 / item_no
- 名称 / name        （必填）
- 规格 / spec
- 单位 / unit
- 工程量 / qty
- 单价 / unit_price
- 分类 / category
- 父级编号 / parent_no
"""
from typing import Any

_HEADER_MAP = {
    "编号": "item_no", "项目编号": "item_no", "item_no": "item_no", "itemno": "item_no",
    "名称": "name", "name": "name",
    "规格": "spec", "spec": "spec",
    "单位": "unit", "unit": "unit",
    "工程量": "qty", "数量": "qty", "qty": "qty", "quantity": "qty",
    "单价": "unit_price", "综合单价": "unit_price", "unit_price": "unit_price", "price": "unit_price",
    "分类": "category", "category": "category", "科目": "category",
    "父级编号": "parent_no", "父编号": "parent_no", "parent_no": "parent_no", "parent": "parent_no",
}


def _norm_header(cell_value: Any) -> str:
    return str(cell_value).strip().lower()


def parse_budget_xlsx(content: bytes) -> list[dict[str, Any]]:
    """解析预算清单 xlsx，返回字段字典列表（已按表头映射）。"""
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        return []
    # 建立 列索引 -> 字段名
    col_map: dict[int, str] = {}
    for idx, h in enumerate(header_cells):
        if h is None:
            continue
        key = _norm_header(h)
        field = _HEADER_MAP.get(key)
        if field:
            col_map[idx] = field

    result: list[dict[str, Any]] = []
    for row in rows_iter:
        rec: dict[str, Any] = {}
        for idx, field in col_map.items():
            val = row[idx] if idx < len(row) else None
            rec[field] = val
        # 整行空则跳过
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in rec.values()):
            continue
        result.append(rec)
    return result
